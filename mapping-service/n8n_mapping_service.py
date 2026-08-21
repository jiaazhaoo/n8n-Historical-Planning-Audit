#!/usr/bin/env python3
"""Small localhost bridge between n8n and the isolated mapping engine.

The service accepts only evidence paths below /data (excluding the protected
file-browser runtime tree), freezes them into an explicit mapping manifest,
runs the existing autonomous capture-rule compiler, and exports two
case-complete mapping tables plus the full audit trail.
"""

from __future__ import annotations

import argparse
import csv
import ipaddress
import json
import re
import shutil
import subprocess
import threading
import traceback
from collections import Counter
from datetime import datetime, timezone
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Iterable

import sys


REPOSITORY_ROOT = Path(__file__).resolve().parent
AUTONOMOUS_ENTRYPOINT = REPOSITORY_ROOT / "amazons3-mapping" / "autonomous_mapping.py"
sys.path.insert(0, str(REPOSITORY_ROOT / "amazons3-mapping"))
from autonomous.path_policy import read_only_job_isolated_command  # noqa: E402
DATA_ROOT = Path("/data")
# Out of bounds for every path this service handles, with no exception left:
# the workflow reads and writes only under /data/<council>/.
PROTECTED_ROOT = DATA_ROOT / "file-browser-data"
JOBS_ROOT = DATA_ROOT / "mapping-jobs"
CODEX_RUNTIME_HOME = JOBS_ROOT / "codex-runtime"
CODEX_AUTH = Path("/home/rmsi/.codex/auth.json")

MAPPING_TOTAL_FIELDS = (
    "batch",
    "originating-authority-charge-identifier",
    "further-information-reference",
    "supplementary-information",
    "charge-address",
    "charge-geographic-description",
    "amazons3_path",
    "amazons3_path_cfd",
    "amazons3_path_mappingrule",
    "amazons3_path_note",
    "portal_path",
    "portal_path_cfd",
    "portal_path_mappingrule",
    "portal_path_note",
    "path_source",
    "path_found",
    # Filled by the download stage, matched on the charge identifier. The
    # mapping stage leaves all four empty: it has downloaded nothing, and a
    # table that writes a local path it never verified starts claiming files
    # that are not there.
    "local_amazons3_path",
    "local_amazons3_file_count",
    "local_portal_path",
    "local_portal_file_count",
)

MAPPINGJ_FIELDS = (
    "batch",
    "originating-authority-charge-identifier",
    "further-information-reference",
    "supplementary-information",
    "charge-address",
    "charge-geographic-description",
    "amazons3_path",
    "portal_path",
    "path_source",
    "path_found",
)

SUPPORTED_TABLES = {".csv", ".json", ".jsonl", ".xlsx"}
SUPPORTED_RULES = {".docx", ".pdf", ".txt"}
REQUEST_LOCK = threading.Lock()


class MappingServiceError(RuntimeError):
    pass


def utc_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def slug(value: str, *, label: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", "-", value.casefold()).strip("-")
    if not cleaned:
        raise MappingServiceError(f"{label} must contain letters or digits")
    return cleaned


def council_matching_table(council: str) -> Path:
    """The council's own matching table, which is the only one this service touches.

    Everything this workflow reads or writes lives under /data/<council>/. The
    file-browser tree holds a copy of the same table, identical in columns and
    row count, and is out of bounds: the service is not the thing that decides
    what a browser serves.
    """
    return DATA_ROOT / council / "file-matching" / f"{council}-matching.csv"


def safe_data_path(value: str, *, kind: str = "file") -> Path:
    if not value or "\x00" in value:
        raise MappingServiceError(f"A nonblank {kind} path is required")
    path = Path(value).expanduser().resolve(strict=True)
    if path != DATA_ROOT and DATA_ROOT not in path.parents:
        raise MappingServiceError(f"{kind} path must be below /data: {path}")
    if path == PROTECTED_ROOT or PROTECTED_ROOT in path.parents:
        raise MappingServiceError("Access to /data/file-browser-data is forbidden")
    if kind == "directory" and not path.is_dir():
        raise MappingServiceError(f"Input directory does not exist: {path}")
    if kind == "file" and not path.is_file():
        raise MappingServiceError(f"Evidence file does not exist: {path}")
    return path


def list_value(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return [part.strip() for part in re.split(r"[\r\n,]+", str(value)) if part.strip()]


def classify_input_files(directory: Path) -> tuple[list[Path], list[Path], list[Path]]:
    files = sorted((path for path in directory.iterdir() if path.is_file()), key=lambda path: path.name.casefold())
    if len(files) > 40:
        raise MappingServiceError("Input directory contains more than 40 files")
    source: list[Path] = []
    rules: list[Path] = []
    breakdown: list[Path] = []
    for path in files:
        words = " ".join(re.findall(r"[a-z0-9]+", path.name.casefold()))
        suffix = path.suffix.casefold()
        if suffix in SUPPORTED_RULES and "capture" in words and "rule" in words:
            rules.append(path)
        elif suffix in SUPPORTED_TABLES and ("source breakdown" in words or "data source breakdown" in words):
            breakdown.append(path)
        elif suffix in SUPPORTED_TABLES and (
            "source information" in words
            or "source record" in words
            or "landmark" in words
            or re.search(r"\b[a-z]{2,5}\s+prod\d+\b", words)
        ):
            source.append(path)
    return source, rules, breakdown


def select_source_table(paths: list[Path]) -> list[Path]:
    """Reduce candidates that are the same table, or a derivative of it.

    Sheffield's Work Package 1 folder holds three files that all read as source
    information: the table, the same table as CSV, and a geocoding-input
    variant. Refusing to choose blocked the run outright, and neither reduction
    below involves a guess about content.

    A file whose stem another file extends with a qualifier is the base table:
    a derivative names what it is for, an original does not. Anything the two
    rules cannot reduce to one is still refused -- picking the wrong table maps
    the wrong population, which is worse than stopping.
    """
    if len(paths) <= 1:
        return paths
    by_stem: dict[str, list[Path]] = {}
    for path in paths:
        by_stem.setdefault(path.stem.casefold().strip(), []).append(path)
    # One table written twice: prefer the spreadsheet, which carries types.
    reduced = [
        sorted(group, key=lambda path: (path.suffix.casefold() != ".xlsx", path.name))[0]
        for group in by_stem.values()
    ]
    stems = {path.stem.casefold().strip(): path for path in reduced}
    bases = [
        path
        for stem, path in stems.items()
        if not any(other != stem and stem.startswith(other) for other in stems)
    ]
    return sorted(bases, key=lambda path: path.name) if bases else sorted(reduced, key=lambda path: path.name)


def select_latest_rule_versions(paths: Iterable[Path]) -> list[Path]:
    """Keep the highest vN within an otherwise identical filename family."""
    selected: dict[str, tuple[int, Path]] = {}
    for path in paths:
        stem = re.sub(r"^\d{8}T\d{6}Z_", "", path.stem, flags=re.IGNORECASE)
        match = re.search(r"(?:^|[-_\s])v(\d+)(?:$|[-_\s])", stem, flags=re.IGNORECASE)
        version = int(match.group(1)) if match else 0
        family = re.sub(r"(?:^|[-_\s])v\d+(?:$|[-_\s])", " ", stem, flags=re.IGNORECASE)
        family = " ".join(re.findall(r"[a-z0-9]+", family.casefold()))
        current = selected.get(family)
        if current is None or version > current[0]:
            selected[family] = (version, path)
    return [item[1] for item in sorted(selected.values(), key=lambda item: item[1].name.casefold())]


def write_csv(path: Path, rows: Iterable[dict[str, str]], fields: tuple[str, ...]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


EXCEL_CELL_LIMIT = 32767


# Control characters are legal in a CSV cell and not in an XLSX one. Mansfield's
# charge text carries DLE bytes inside a condition -- "...2021/0340/TPO \x10TOLD
# TO SUBMIT DISCHARGE OF CONDITIONS APP\x10..." -- and openpyxl refused the
# whole export, so a mapping that had completed produced no spreadsheet.
ILLEGAL_XLSX_CHARACTERS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def xlsx_text(value: str) -> str:
    """The cell text, with characters a spreadsheet cannot hold removed.

    Dropped rather than escaped: these are stray bytes in scanned-record text,
    not content, and the CSV beside it keeps the value exactly as it was.
    """
    return ILLEGAL_XLSX_CHARACTERS.sub("", value)


def write_xlsx(path: Path, rows: Iterable[dict[str, str]], fields: tuple[str, ...], *, sheet: str) -> None:
    """Write a mapping table as .xlsx alongside the canonical CSV.

    Every cell is written as text. Charge identifiers, references and paths look
    numeric or date-like often enough that letting Excel infer a type silently
    rewrites the value the mapping is supposed to preserve.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise MappingServiceError(
            "openpyxl is required to export .xlsx mapping tables"
        ) from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet[:31] or "mapping"
    worksheet.append(list(fields))
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    widths = [len(field) for field in fields]
    for row in rows:
        values = []
        for index, field in enumerate(fields):
            value = xlsx_text(str(row.get(field, "") or ""))
            if len(value) > EXCEL_CELL_LIMIT:
                raise MappingServiceError(
                    f"Value for {field!r} exceeds the Excel cell limit; export the CSV instead"
                )
            widths[index] = max(widths[index], min(len(value), 60))
            values.append(value)
        worksheet.append(values)
        for cell in worksheet[worksheet.max_row]:
            # Leading "=", "+" or "-" would otherwise be stored as a formula.
            cell.data_type = "s"
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = width + 2
    worksheet.freeze_panes = "A2"
    workbook.save(path)


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="", errors="replace") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def source_value(row: dict[str, str], field: str) -> str:
    """Read a requested Landmark field without rewriting its source value."""
    if field in row:
        return (row.get(field) or "").strip()
    normalized = re.sub(r"[^a-z0-9]+", "", field.casefold())
    for key, value in row.items():
        if re.sub(r"[^a-z0-9]+", "", key.casefold()) == normalized:
            return (value or "").strip()
    return ""


def export_outputs(*, workspace: Path, council: str, batch: str, output_directory: Path) -> dict[str, Any]:
    mapping_path = workspace / "mapping" / "proposed-mapping.csv"
    audit_path = workspace / "mapping" / "mapping-audit.csv"
    spec_path = workspace / "spec" / "mapping-spec.json"
    validation_path = workspace / "validation" / "validation.json"
    source_path = workspace / "prepared" / "source-records.csv"
    for path in (mapping_path, audit_path, spec_path, validation_path, source_path):
        if not path.is_file():
            raise MappingServiceError(f"Mapping job did not produce required artifact: {path}")

    mappings = read_csv(mapping_path)
    audits = read_csv(audit_path)
    source_rows = read_csv(source_path)
    spec = json.loads(spec_path.read_text(encoding="utf-8"))
    source_id_field = str(spec.get("source_id_field") or "oachargeid")
    if len(mappings) != len(audits):
        raise MappingServiceError("Mapping and audit row counts differ")
    audit_by_id = {row.get("oachargeid", ""): row for row in audits}
    if len(audit_by_id) != len(audits):
        raise MappingServiceError("Mapping audit contains blank or duplicate case identifiers")
    source_by_id = {source_value(row, source_id_field): row for row in source_rows}
    if len(source_by_id) != len(source_rows) or any(not oid for oid in source_by_id):
        raise MappingServiceError("Source evidence contains blank or duplicate case identifiers")

    mapping_total_rows: list[dict[str, str]] = []
    mappingj_rows: list[dict[str, str]] = []
    for mapping in mappings:
        oid = (mapping.get("oachargeid") or "").strip()
        audit = audit_by_id.get(oid)
        source = source_by_id.get(oid)
        if audit is None:
            raise MappingServiceError(f"Missing audit row for case {oid!r}")
        if source is None:
            raise MappingServiceError(f"Missing source row for case {oid!r}")
        raw_status = (audit.get("match_status") or "").strip()
        if raw_status.startswith("accepted_"):
            status = "found"
        elif raw_status == "not_found":
            status = "no_found"
        else:
            status = "no_match"
        route = (audit.get("route") or "").strip()
        source_type = "amazons3" if route == "s3" else ("portal" if route == "portal" else "")
        reason = (audit.get("rejection_reason") or "").strip() or raw_status
        amazon_path = (mapping.get("amazons3_path") or "").strip()
        portal_path = (mapping.get("portal_path") or "").strip()
        row = {
                "batch": batch,
                "originating-authority-charge-identifier": oid,
                "further-information-reference": source_value(source, "further-information-reference"),
                "supplementary-information": source_value(source, "supplementary-information"),
                "charge-address": source_value(source, "charge-address"),
                "charge-geographic-description": source_value(source, "charge-geographic-description"),
                "amazons3_path": amazon_path,
                "amazons3_path_cfd": (
                    (mapping.get("amazons3_confidence") or "0.00").strip() if route == "s3" else "0.00"
                ),
                "amazons3_path_mappingrule": (audit.get("rule_id") or "").strip() if route == "s3" else "",
                "amazons3_path_note": reason if route == "s3" else "",
                "portal_path": portal_path,
                "portal_path_cfd": (
                    (audit.get("mapping_confidence") or audit.get("decision_confidence") or "0.00").strip() if route == "portal" else "0.00"
                ),
                "portal_path_mappingrule": (audit.get("rule_id") or "").strip() if route == "portal" else "",
                "portal_path_note": reason if route == "portal" else "",
                "path_source": source_type,
                "path_found": status,
            }
        mapping_total_rows.append(row)
        mappingj_rows.append({field: row[field] for field in MAPPINGJ_FIELDS})

    ids = [row["originating-authority-charge-identifier"] for row in mapping_total_rows]
    if any(not oid for oid in ids) or len(ids) != len(set(ids)):
        raise MappingServiceError("Output contains blank or duplicate case identifiers")

    output_directory.mkdir(parents=True, exist_ok=False)
    prefix = f"{council}-{slug(batch, label='batch')}"
    # A batch has one current mapping, so the three delivered tables sit at a
    # fixed path in the council's own folder rather than inside this run's
    # timestamped directory. Exeter's wp3 accumulated ten run directories, each
    # with its own table, which left the download stage no answer to which one
    # it should fill its four columns into.
    delivery = DATA_ROOT / council / "file-matching"
    delivery.mkdir(parents=True, exist_ok=True)
    mapping_total_path = delivery / f"{prefix}-full-mapping.csv"
    mappingj_path = delivery / f"{prefix}-mapping.csv"
    exported_audit_path = delivery / f"{prefix}-mapping-audit.csv"
    full_mapping_xlsx_path = delivery / f"{prefix}-full-mapping.xlsx"
    mappingj_xlsx_path = delivery / f"{prefix}-mapping.xlsx"
    # The spec and the report describe this run, not the batch, so they stay
    # with the run that produced them.
    exported_spec_path = output_directory / "routing-spec.json"
    report_path = output_directory / "mapping-run-report.json"
    write_csv(mapping_total_path, mapping_total_rows, MAPPING_TOTAL_FIELDS)
    write_csv(mappingj_path, mappingj_rows, MAPPINGJ_FIELDS)
    # The CSVs stay the runtime contract that file-browser reads; the workbooks
    # are the human-facing deliverable and must not diverge from them.
    write_xlsx(full_mapping_xlsx_path, mapping_total_rows, MAPPING_TOTAL_FIELDS, sheet="full mapping")
    write_xlsx(mappingj_xlsx_path, mappingj_rows, MAPPINGJ_FIELDS, sheet="mapping")
    shutil.copy2(audit_path, exported_audit_path)
    shutil.copy2(spec_path, exported_spec_path)
    validation = json.loads(validation_path.read_text(encoding="utf-8"))
    counts = Counter(row["path_found"] for row in mapping_total_rows)
    source_counts = Counter(row["path_source"] or "none" for row in mapping_total_rows)
    report = {
        "status": "completed_staged",
        "production_published": False,
        "council": council,
        "batch": batch,
        "case_count": len(mapping_total_rows),
        "match_status_counts": {key: counts.get(key, 0) for key in ("found", "no_found", "no_match")},
        "source_type_counts": dict(sorted(source_counts.items())),
        "coverage": {
            "mapping_rows": len(mappings),
            "audit_rows": len(audits),
            "unique_case_ids": len(set(ids)),
            "exact": len(mappings) == len(audits) == len(set(ids)),
        },
        "validation": validation,
        "outputs": {
            "mapping_total": str(mapping_total_path),
            "mappingj": str(mappingj_path),
            "full_mapping_xlsx": str(full_mapping_xlsx_path),
            "mappingj_xlsx": str(mappingj_xlsx_path),
            "audit": str(exported_audit_path),
            "routing_spec": str(exported_spec_path),
            "run_report": str(report_path),
        },
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return report


def copy_artifact(path: Path, staging: Path, index: int) -> str:
    destination = staging / f"{index:02d}_{path.name}"
    shutil.copy2(path, destination)
    return destination.name


INVENTORY_SOURCES_NAME = "inventory-sources.json"
INVENTORY_NAME_SUFFIX = "-s3-index.csv"


def _s3_lister(bucket: str, prefix: str):
    import boto3
    from key_manager import get_aws_client_kwargs

    client = boto3.client("s3", **get_aws_client_kwargs())
    paginator = client.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        yield from page.get("Contents", [])


def ensure_inventory(council: str, *, rebuild: bool = False) -> tuple[Path | None, dict[str, Any]]:
    """Return the council's scan inventory, building it once if it is absent.

    The inventory does not vary by batch and listing one can take minutes, so it
    is built on first use and reused after. Which trees to list, and whether a
    record is a folder or a file, are declared per council rather than asked for
    on every run: getting them wrong is silent, and a run should not be able to
    get them wrong differently each time.
    """
    from autonomous.inventory import (
        InventoryError,
        build_rows,
        load_sources,
        missing_trees,
        summarise,
        write_inventory,
    )

    root = DATA_ROOT / council / "file-matching"
    declaration = root / INVENTORY_SOURCES_NAME
    inventory = root / f"{council}{INVENTORY_NAME_SUFFIX}"
    if not declaration.is_file():
        return None, {"declared": False, "reason": f"no {declaration}"}
    if inventory.is_file() and not rebuild:
        return inventory, {"declared": True, "built": False, "path": str(inventory)}

    try:
        sources = load_sources(declaration)
        rows = build_rows(sources, lister=_s3_lister)
    except InventoryError as exc:
        raise MappingServiceError(str(exc)) from exc

    summary = summarise(rows, sources)
    delivered_csv = council_matching_table(council)
    if delivered_csv.is_file():
        # A tree the delivered mapping points into but no source lists is a
        # missing tree, and it is far cheaper to say so here than to infer it
        # later from an unexplained join rate.
        delivered = [row.get("amazons3_path", "") for row in read_csv(delivered_csv)]
        missing = missing_trees(sources, delivered)
        summary["missing_trees"] = missing
        if missing:
            raise MappingServiceError(
                "The declared inventory sources do not cover trees the delivered mapping uses: "
                f"{missing}. Add them to {declaration} before mapping."
            )

    write_inventory(rows, inventory)
    (root / f"{council}-s3-index-summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    summary.update({"declared": True, "built": True, "path": str(inventory)})
    return inventory, summary


MAX_PRIOR_FINDINGS = 12
MAX_FINDING_CHARACTERS = 600


def prior_findings_value(value: Any) -> list[list[str]]:
    """Normalise what a caller says an earlier spec got wrong.

    Bounded on both axes: the prompt has to stay readable, and a caller looping
    a failing batch must not grow it without limit.
    """
    if not value:
        return []
    rounds = value if isinstance(value, list) else [value]
    findings: list[list[str]] = []
    for entry in rounds[:MAX_PRIOR_FINDINGS]:
        items = entry if isinstance(entry, list) else [entry]
        cleaned = [
            str(item).strip()[:MAX_FINDING_CHARACTERS]
            for item in items
            if item is not None and str(item).strip()
        ]
        if cleaned:
            findings.append(cleaned)
    return findings


PORTAL_EVIDENCE_NAME = "portal-evidence.json"
PORTAL_URL = re.compile(r"https?://[^\s\]<>\"']{6,200}")
PORTAL_WORDS = re.compile(r"public access|planning portal|online-applications", re.I)


def portal_named_in_rules(rule_paths: Iterable[Path]) -> str | None:
    """The portal a council's capture rules point at, if they point at one.

    The rules name it: Test Valley's say the Public Access Planning Portal at
    view-applications.testvalley.gov.uk covers 2005 onward. What they cannot say
    is which case is which document -- that only a search of the portal
    establishes, and path construction is forbidden -- so naming it is how a run
    reports that the evidence it would need was never supplied.
    """
    for path in rule_paths:
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
        except OSError:
            continue
        for match in PORTAL_URL.finditer(text):
            url = match.group(0).rstrip(".,;")
            window = text[max(0, match.start() - 300) : match.end() + 120]
            if PORTAL_WORDS.search(window) or PORTAL_WORDS.search(url):
                return url
    return None


def declared_portal_evidence(council: str) -> list[Path]:
    """Portal evidence this council has declared, the way S3 trees are declared.

    An S3 inventory is found three ways -- passed in, at a conventional name, or
    built from inventory-sources.json -- and portal evidence only one, by hand.
    Test Valley wp2 lost 17 of its 30 cases to that: the compiler read the rules
    correctly, split the batch at 2004, and routed the later records to a reject
    named reject_portal_period_without_portal_inventory. Nobody was going to
    read a route name.
    """
    declaration = DATA_ROOT / council / "file-matching" / PORTAL_EVIDENCE_NAME
    if not declaration.is_file():
        return []
    try:
        payload = json.loads(declaration.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise MappingServiceError(f"{declaration} is not readable JSON: {exc}") from exc
    entries = payload.get("evidence") if isinstance(payload, dict) else payload
    paths: list[Path] = []
    for entry in entries or []:
        value = str(entry).strip()
        if not value:
            continue
        candidate = Path(value)
        if not candidate.is_absolute():
            candidate = DATA_ROOT / council / "file-matching" / value
        if not candidate.is_file():
            raise MappingServiceError(
                f"{declaration} names portal evidence that does not exist: {candidate}"
            )
        paths.append(safe_data_path(str(candidate)))
    return paths


def run_mapping(payload: dict[str, Any]) -> dict[str, Any]:
    council = slug(str(payload.get("council") or ""), label="council")
    batch = str(payload.get("batch") or "").strip()
    if not batch:
        raise MappingServiceError("batch is required and must be a registered file-matching batch")
    input_directory = safe_data_path(str(payload.get("input_directory") or ""), kind="directory")

    discovered_source, discovered_rules, discovered_breakdown = classify_input_files(input_directory)
    explicit_source = list_value(payload.get("source_path"))
    source_paths = (
        [safe_data_path(value) for value in explicit_source]
        if explicit_source
        else select_source_table(discovered_source)
    )
    if len(source_paths) != 1:
        listed = ", ".join(sorted(path.name for path in source_paths)) or "none"
        raise MappingServiceError(
            f"Expected exactly one source table; found {len(source_paths)}: {listed}. "
            "Supply source_path explicitly."
        )
    explicit_rules = list_value(payload.get("capture_rules_paths"))
    rule_paths = [safe_data_path(value) for value in explicit_rules] if explicit_rules else select_latest_rule_versions(discovered_rules)
    if not rule_paths:
        raise MappingServiceError("No capture-rules DOCX, PDF, or TXT file was found")

    s3_paths = [safe_data_path(value) for value in list_value(payload.get("s3_inventory_paths"))]
    portal_paths = [safe_data_path(value) for value in list_value(payload.get("portal_evidence_paths"))]
    inventory_summary: dict[str, Any] = {"declared": False}
    if not s3_paths and not portal_paths:
        conventional = DATA_ROOT / council / "file-matching" / f"{council}-s3-folder-index.csv"
        if conventional.is_file():
            s3_paths = [safe_data_path(str(conventional))]
    if not s3_paths and not portal_paths:
        built, inventory_summary = ensure_inventory(
            council, rebuild=bool(payload.get("rebuild_inventory", False))
        )
        if built is not None:
            s3_paths = [safe_data_path(str(built))]
    # Last, and only when no scan inventory answered. Portal records are their
    # own batch -- testvalley declares one, braintree two, torbay nothing else
    # -- so a scan batch that found its inventory is complete, and reaching for
    # portal evidence would blur a boundary the registry keeps on purpose.
    if not s3_paths and not portal_paths:
        portal_paths = declared_portal_evidence(council)
    if not s3_paths and not portal_paths:
        raise MappingServiceError(
            "No S3 inventory or Portal evidence was supplied, and this council declares neither. "
            f"Create {DATA_ROOT / council / 'file-matching' / INVENTORY_SOURCES_NAME} for scans, "
            f"{DATA_ROOT / council / 'file-matching' / PORTAL_EVIDENCE_NAME} for portal records, "
            "or pass the paths explicitly."
        )

    stamp = utc_stamp()
    submission = JOBS_ROOT / "submissions" / f"{council}-{slug(batch, label='batch')}-{stamp}"
    submission.mkdir(parents=True, exist_ok=False)
    artifacts: list[dict[str, str]] = []
    index = 1
    for role, paths in (
        ("source_records", source_paths),
        ("capture_rules", rule_paths),
        ("source_breakdown", discovered_breakdown),
        ("s3_inventory", s3_paths),
        ("portal_evidence", portal_paths),
    ):
        for path in paths:
            artifacts.append({"url": copy_artifact(path, submission, index), "role": role})
            index += 1
    manifest_path = submission / "mapping-job.json"
    manifest_path.write_text(
        json.dumps(
            {
                "kind": "council-mapping-job",
                "schema_version": 1,
                "council": council,
                "batch": batch,
                "artifacts": artifacts,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    approved_spec_value = str(payload.get("mapping_spec_path") or "").strip()
    approved_spec = safe_data_path(approved_spec_value) if approved_spec_value else None
    if approved_spec is None:
        conventional_spec = (
            DATA_ROOT / council / "file-matching" / "approved-specs" / f"{slug(batch, label='batch')}.json"
        )
        if conventional_spec.is_file():
            # A spec already verified for this batch is reused by default, so a
            # working mapping does not depend on the compiler getting it right
            # again on every run.
            approved_spec = safe_data_path(str(conventional_spec))

    command = [
        str(AUTONOMOUS_ENTRYPOINT),
        "--jobs-root",
        str(JOBS_ROOT),
        "start",
        "--url",
        manifest_path.as_uri(),
        "--council",
        council,
        "--batch",
        batch,
        "--requested-by",
        "n8n-file-path-mapping",
    ]
    if approved_spec is not None:
        command.extend(("--approved-spec", str(approved_spec)))

    # A rework carries what the last quality round found wrong. Written beside
    # the submission so the reason a spec was rewritten stays with the job that
    # rewrote it.
    findings = prior_findings_value(payload.get("prior_findings"))
    if findings:
        findings_path = submission / "prior-findings.json"
        findings_path.write_text(
            json.dumps(findings, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
        command.extend(("--prior-findings", str(findings_path)))
    result = subprocess.run(
        read_only_job_isolated_command(
            command,
            writable_root=JOBS_ROOT,
            codex_home=CODEX_RUNTIME_HOME,
            codex_auth=CODEX_AUTH,
        ),
        cwd=REPOSITORY_ROOT,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        timeout=1800,
        check=False,
    )
    try:
        snapshot = json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise MappingServiceError(
            f"Mapping engine returned unreadable output (exit {result.returncode}): {(result.stderr or result.stdout)[-2000:]}"
        ) from exc
    job_record = snapshot.get("job") if isinstance(snapshot.get("job"), dict) else snapshot
    if result.returncode != 0 or job_record.get("status") not in {
        "completed_staged",
        "completed_safe_with_exceptions",
    }:
        detail = (
            job_record.get("error")
            or snapshot.get("error")
            or snapshot.get("detail")
            or job_record.get("status")
            or "mapping failed"
        )
        raise MappingServiceError(f"Mapping engine stopped: {detail}")
    workspace = safe_data_path(str(job_record.get("workspace") or ""), kind="directory")
    output_directory = DATA_ROOT / council / "file-matching" / f"{slug(batch, label='batch')}_{stamp}"
    report = export_outputs(
        workspace=workspace,
        council=council,
        batch=batch,
        output_directory=output_directory,
    )
    report["job_id"] = job_record.get("job_id")
    report["input_directory"] = str(input_directory)
    report["selected_source"] = [str(path) for path in source_paths]
    report["selected_capture_rules"] = [str(path) for path in rule_paths]
    report["selected_s3_inventory"] = [str(path) for path in s3_paths]
    report["selected_portal_evidence"] = [str(path) for path in portal_paths]
    report["approved_spec"] = str(approved_spec) if approved_spec else None
    report["spec_source"] = "approved" if approved_spec else "compiled"
    # Echoed so a rework loop reads its own depth from the data rather than from
    # a node that the loop's return path never re-executes.
    report["rework_round"] = int(payload.get("rework_round") or 0)
    report["prior_findings"] = findings
    report["inventory"] = inventory_summary
    # A batch that needs the portal and was not given it rejects those rows
    # through a capture rule, which reads like a decision rather than a missing
    # input. Test Valley wp2 lost 17 of its 30 cases that way without anything
    # in the result saying so.
    report["missing_evidence"] = []
    if not portal_paths:
        portal = portal_named_in_rules(rule_paths)
        rejected = int((report.get("match_status_counts") or {}).get("no_match", 0))
        if portal and rejected:
            report["missing_evidence"].append(
                {
                    "kind": "portal",
                    "named_in_capture_rules": portal,
                    "unrouted_cases": rejected,
                    "detail": (
                        f"The capture rules name a planning portal at {portal}, no portal evidence "
                        f"was supplied, and {rejected} case(s) reached no accepting route. Portal "
                        "paths cannot be constructed from a reference, so the portal has to be "
                        "searched and its results supplied as evidence -- pass "
                        f"portal_evidence_paths, or declare them in {DATA_ROOT / council / 'file-matching' / PORTAL_EVIDENCE_NAME}."
                    ),
                }
            )
    # export_outputs wrote the report before these run-level fields existed. The
    # quality loop reads the report from disk to find the audit and source it
    # must review, so the file has to carry what the response carries.
    Path(report["outputs"]["run_report"]).write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    return report


QUALITY_LOOP_DIRNAME = "quality-loop"
# Each sampling round is priced independently, so a loop that needs three rounds
# costs three budgets rather than dividing one.
DEFAULT_QA_MODEL = "google/gemini-3.7-flash"
# Judging fit is a smaller task than extracting identity, so it runs on a
# cheaper model: measured $0.00035 a case.
DEFAULT_QA_JUDGE_MODEL = "deepseek/deepseek-v4-flash"
DEFAULT_QA_BUDGET_USD = 3.0
DEFAULT_QA_ESTIMATE_USD = 0.01
QUALITY_REPORT_DIRNAME = "quality-report"


def quality_state_path(council: str, batch: str) -> Path:
    """One loop per council and batch, so the split survives every re-run."""
    return (
        DATA_ROOT
        / council
        / "file-matching"
        / QUALITY_LOOP_DIRNAME
        / f"{slug(batch, label='batch')}.json"
    )


def load_run_report(run_directory: Path) -> dict[str, Any]:
    report_path = run_directory / "mapping-run-report.json"
    if not report_path.is_file():
        raise MappingServiceError(
            f"{run_directory} is not a mapping run directory; mapping-run-report.json is absent"
        )
    report = json.loads(report_path.read_text(encoding="utf-8"))
    if not isinstance(report, dict):
        raise MappingServiceError("mapping-run-report.json must contain an object")
    return report


def run_quality(payload: dict[str, Any]) -> dict[str, Any]:
    """Review one sample of a completed mapping run and judge it.

    A working round tests cases the spec may still be adjusted against; the
    acceptance round tests the reserved holdout and is the only round whose
    result describes the mapping rather than the loop.
    """
    from autonomous.content_qa import IdentityFieldProfile
    from autonomous.content_reviewer import ContentQaReviewer, ReviewerSettings
    from dataclasses import replace

    from autonomous.quality_gate import GateThresholds, evaluate_coverage
    from autonomous.quality_loop import (
        ACCEPTANCE,
        WORKING,
        load_audit_rows,
        load_round_records,
        next_action,
        open_state,
        run_quality_round,
        spent_so_far,
        write_report,
    )

    council = slug(str(payload.get("council") or ""), label="council")
    batch = str(payload.get("batch") or "").strip()
    if not batch:
        raise MappingServiceError("batch is required")
    run_directory = safe_data_path(str(payload.get("run_directory") or ""), kind="directory")
    report = load_run_report(run_directory)

    stage = str(payload.get("stage") or WORKING).strip().lower()
    if stage not in {WORKING, ACCEPTANCE}:
        raise MappingServiceError(f"stage must be {WORKING!r} or {ACCEPTANCE!r}")

    outputs = report.get("outputs") or {}
    audit_path = safe_data_path(str(outputs.get("audit") or ""))
    explicit_source = list_value(payload.get("source_path"))
    source_candidates = explicit_source or list_value(report.get("selected_source"))
    if len(source_candidates) != 1:
        raise MappingServiceError(
            "Exactly one source table is required; supply source_path explicitly"
        )
    source_path = safe_data_path(source_candidates[0])

    documents_root_value = str(payload.get("documents_root") or "").strip()
    documents_root = safe_data_path(documents_root_value, kind="directory") if documents_root_value else None
    acquire = bool(payload.get("acquire", documents_root is None))

    settings = ReviewerSettings(
        council=council,
        batch=batch,
        audit_path=audit_path,
        source_path=source_path,
        source_id_field=str(
            payload.get("source_id_field") or "originating-authority-charge-identifier"
        ),
        source_original_name=str(payload.get("source_original_name") or "").strip(),
        documents_root=documents_root,
        acquire=acquire,
        max_images_per_case=int(payload.get("max_images", 12)),
        max_image_pixels=int(payload.get("max_image_pixels", 80_000_000)),
        extractor_mode=str(payload.get("extractor_mode") or "judge").strip().lower(),
        judge_model=str(payload.get("judge_model") or "").strip() or DEFAULT_QA_JUDGE_MODEL,
        model=str(payload.get("model") or "").strip() or DEFAULT_QA_MODEL,
        budget_usd=float(payload.get("budget_usd", DEFAULT_QA_BUDGET_USD)),
        estimate_usd_per_case=float(
            payload.get("estimate_usd_per_case", DEFAULT_QA_ESTIMATE_USD)
        ),
        field_profile=IdentityFieldProfile(
            reference_fields=tuple(list_value(payload.get("reference_field"))),
            address_fields=tuple(list_value(payload.get("address_field"))),
            description_fields=tuple(list_value(payload.get("description_field"))),
            date_fields=tuple(list_value(payload.get("date_field"))),
            document_type_fields=tuple(list_value(payload.get("document_type_field"))),
        ),
    )

    thresholds = GateThresholds(
        min_verified_rate=float(payload.get("min_verified_rate", 0.80)),
        max_verified_wrong=int(payload.get("max_verified_wrong", 0)),
        max_systematic_failures=int(payload.get("max_systematic_failures", 0)),
        max_missing_document_rate=float(payload.get("max_missing_document_rate", 0.10)),
    )
    thresholds.validate()

    audit_rows = load_audit_rows(audit_path)
    state_path = quality_state_path(council, batch)
    state = open_state(
        state_path,
        council=council,
        batch=batch,
        audit_rows=audit_rows,
        holdout_fraction=float(payload.get("holdout_fraction", 0.2)),
    )

    # One ceiling for the whole mapping, not one per round. The budget object is
    # rebuilt on every /quality call, and the workflow loops -- up to two
    # recompiles, each drawing up to four sampling rounds -- so a per-round limit
    # multiplies into something nobody asked for.
    already = spent_so_far(state)
    remaining = settings.budget_usd - already
    if remaining <= 0:
        raise MappingServiceError(
            f"This mapping has already spent ${already:.4f} of its ${settings.budget_usd:.2f} "
            "review budget across its rounds. Raise budget_usd to continue, or read the rounds "
            "already reviewed."
        )
    settings = replace(settings, budget_usd=min(settings.budget_usd, remaining))

    result = run_quality_round(
        state=state,
        state_path=state_path,
        audit_rows=audit_rows,
        reviewer=ContentQaReviewer(settings, repository_root=REPOSITORY_ROOT),
        artifacts_root=run_directory / "quality-rounds",
        stage=stage,
        sample_size=int(payload.get("sample_size", 12)),
        thresholds=thresholds,
    )

    # Sample precision and population coverage answer different questions and
    # are judged separately: a spec can be right about everything it accepted
    # while accepting far too little to deliver.
    coverage = evaluate_coverage(
        report.get("match_status_counts"),
        min_accepted_rate=float(payload.get("min_accepted_rate", 0.0)),
        max_unmatched_rate=float(payload.get("max_unmatched_rate", 1.0)),
    )
    action = next_action(result)
    if action.get("action") in {"accept", "publish"} and not coverage.passed:
        action = {
            "action": "adjust_spec",
            "detail": (
                "The reviewed sample was accurate, but the mapping did not resolve enough of "
                "the population to publish."
            ),
            "focus": list(coverage.reasons),
        }

    response: dict[str, Any] = {
        "council": council,
        "batch": batch,
        "run_directory": str(run_directory),
        "production_published": False,
        "state_path": str(state_path),
        "sampling_plan": state.plan().describe(),
        "budget": result.record.verification_report.get("budget"),
        "model": result.record.verification_report.get("model"),
        "extractor_mode": result.record.verification_report.get("extractor_mode"),
        **result.describe(),
        "coverage": coverage.describe(),
        "passed": result.outcome.passed and coverage.passed,
        "next": action,
    }

    if bool(payload.get("write_report", True)) and not result.exhausted:
        rounds, acceptance = load_round_records(state)
        destination = write_report(
            destination=run_directory / QUALITY_REPORT_DIRNAME,
            state=state,
            rounds=rounds,
            acceptance=acceptance,
            mapping_summary={
                "case_count": report.get("case_count"),
                "coverage": coverage.describe(),
                "match_status_counts": report.get("match_status_counts"),
                "source_type_counts": report.get("source_type_counts"),
                "outputs": outputs,
            },
        )
        response["quality_report"] = {
            "directory": str(destination),
            "html": str(destination / "index.html"),
            "summary": str(destination / "quality-report.json"),
        }
    return response


def run_publish(payload: dict[str, Any]) -> dict[str, Any]:
    """Promote one verified run into the runtime table it serves from.

    This is the only path that writes below /data/file-browser-data. Every other
    entry point refuses that tree outright, and this one opens only on the run's
    own acceptance evidence.
    """
    from autonomous.publication import PublicationError, publish_mapping

    council = slug(str(payload.get("council") or ""), label="council")
    batch = str(payload.get("batch") or "").strip()
    if not batch:
        raise MappingServiceError("batch is required")
    run_directory = safe_data_path(str(payload.get("run_directory") or ""), kind="directory")
    report = load_run_report(run_directory)
    if report.get("council") != council or report.get("batch") != batch:
        raise MappingServiceError(
            f"Run directory belongs to {report.get('council')}/{report.get('batch')}, "
            f"not {council}/{batch}"
        )

    mapping_csv = safe_data_path(str((report.get("outputs") or {}).get("mapping_total") or ""))
    runtime_csv = council_matching_table(council)
    if not runtime_csv.is_file():
        raise MappingServiceError(f"Council matching table does not exist: {runtime_csv}")
    # Routed through the same guard as everything else, so no path in this
    # service can reach the file-browser tree even by construction.
    runtime_csv = safe_data_path(str(runtime_csv))

    try:
        result = publish_mapping(
            council=council,
            batch=batch,
            mapping_csv=mapping_csv,
            runtime_csv=runtime_csv,
            run_directory=run_directory,
            dry_run=bool(payload.get("dry_run", False)),
        )
    except PublicationError as exc:
        raise MappingServiceError(str(exc)) from exc

    described = result.describe()
    described["dry_run"] = bool(payload.get("dry_run", False))
    if described["dry_run"]:
        described["production_published"] = False
    return described


class Handler(BaseHTTPRequestHandler):
    server_version = "n8n-file-path-mapping/1.0"

    def log_message(self, format: str, *args: object) -> None:
        print(f"{self.address_string()} - {format % args}", flush=True)

    def respond(self, status: int, payload: dict[str, Any]) -> None:
        body = (json.dumps(payload, ensure_ascii=False, indent=2) + "\n").encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def client_allowed(self) -> bool:
        address = ipaddress.ip_address(self.client_address[0])
        return address.is_loopback or address in ipaddress.ip_network("172.16.0.0/12")

    def do_GET(self) -> None:  # noqa: N802
        if not self.client_allowed():
            self.respond(HTTPStatus.FORBIDDEN, {"error": "forbidden_client"})
            return
        if self.path == "/healthz":
            self.respond(HTTPStatus.OK, {"status": "ok", "service": "file-path-mapping"})
            return
        self.respond(HTTPStatus.NOT_FOUND, {"error": "not_found"})

    def do_POST(self) -> None:  # noqa: N802
        if not self.client_allowed():
            self.respond(HTTPStatus.FORBIDDEN, {"error": "forbidden_client"})
            return
        handlers = {"/run": run_mapping, "/quality": run_quality, "/publish": run_publish}
        handler = handlers.get(self.path)
        if handler is None:
            self.respond(HTTPStatus.NOT_FOUND, {"error": "not_found"})
            return
        try:
            length = int(self.headers.get("Content-Length", "0"))
            if length <= 0 or length > 1024 * 1024:
                raise MappingServiceError("Request body must be JSON and no larger than 1 MiB")
            payload = json.loads(self.rfile.read(length))
            if not isinstance(payload, dict):
                raise MappingServiceError("Request JSON must be an object")
            if not REQUEST_LOCK.acquire(blocking=False):
                self.respond(HTTPStatus.CONFLICT, {"error": "mapping_job_already_running"})
                return
            try:
                result = handler(payload)
            finally:
                REQUEST_LOCK.release()
            self.respond(HTTPStatus.OK, result)
        except (RuntimeError, OSError, ValueError, subprocess.TimeoutExpired) as exc:
            self.respond(
                HTTPStatus.UNPROCESSABLE_ENTITY,
                {"error": type(exc).__name__, "detail": str(exc), "production_published": False},
            )
        except Exception as exc:  # pragma: no cover - last-resort service boundary
            traceback.print_exc()
            self.respond(
                HTTPStatus.INTERNAL_SERVER_ERROR,
                {"error": type(exc).__name__, "detail": str(exc), "production_published": False},
            )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=5680)
    args = parser.parse_args()
    server = ThreadingHTTPServer((args.host, args.port), Handler)
    print(f"Listening on http://{args.host}:{args.port}", flush=True)
    server.serve_forever()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
